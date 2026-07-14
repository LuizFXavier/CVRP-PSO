import subprocess
import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook, Workbook
from sys import argv
from leitor_instancia import ler_instancia

NUM_TESTES = 5
instance = ""

arquivos_teste = []
caminho = ''

if argv[1].upper() == "-F":
    arquivos_teste.append(argv[2])
elif argv[1].upper() == "-D":
    arquivos_teste = os.listdir(argv[2])
    caminho = argv[2]
    if caminho[-1] != '/':
        caminho += '/'
else:
    raise Exception("Não se sabe se é arquivo ou diretório!")

caminho_output = "./PSO_data.xlsx"

if len(argv) == 4:
    caminho_output = argv[3]

arquivos_teste.sort()

try:
    wb = load_workbook(caminho_output)
    ws = wb.active
    ws.append([])
except:
    wb = Workbook()
    ws = wb.active

dados = [["Mínimo:"],
         ["Mediana:"],
         ["Máximo:"],
         ["Média:"],
         ["Desvio padrão:"],
         ["Tempo Médio (s):"],
         ["SD Tempo:"]]

valores_custo = []
valores_tempo = []

nParticulas = [25]
nRepeticoes = [5000]
nElite = [5]

NUM_THREADS = ["1", "16"]

count = 1

EXECUTAVEL = os.path.expanduser("~/Projects/lscad/CVRP-PSO/src_new/build/default/app/cvrp-pso")

for caso_teste in arquivos_teste:
    print(caso_teste)
    if caso_teste[-3:] != "vrp":
        continue

    instancia = ler_instancia(caminho + caso_teste)

    for num_threads in NUM_THREADS:
      for c in range(len(nRepeticoes)):
          
          print(caso_teste.split("/")[-1], ", threads:", num_threads, ", ", c)
          
          comando = [
              EXECUTAVEL,
              "--in", caminho + caso_teste,
              "--swarm", str(nParticulas[c]),
              "--iter", str(nRepeticoes[c]),
              "--elite", str(nElite[c]),
              "--runs", str(NUM_TESTES),
              "--sector", str(nParticulas[c])
          ]

          environ = os.environ.copy()

          environ ["OMP_NUM_THREADS"] = num_threads
          
          # Captura todas as execuções de uma vez
          output = subprocess.check_output(comando).decode().strip()
          
          # Separa a string por quebras de linha
          linhas_output = output.split('\n')
          
          for idx, linha in enumerate(linhas_output):
              if not linha.strip():
                  continue
                  
              # Divide os dois valores impressos pelo C++
              custo_str, tempo_str = linha.split(',')
              
              print(f"Run {idx + 1} -> Custo: {custo_str} | Tempo: {tempo_str}s")
              
              valores_custo.append(float(custo_str))
              valores_tempo.append(float(tempo_str))
          
          # Consolidação dos dados matemáticos
          dados[0].append(min(valores_custo))
          dados[1].append(np.median(valores_custo))
          dados[2].append(max(valores_custo))
          dados[3].append(np.average(valores_custo))
          dados[4].append(np.std(valores_custo))
          dados[5].append(np.average(valores_tempo))
          dados[6].append(np.std(valores_tempo))

          # Reset para a próxima configuração
          valores_custo = []
          valores_tempo = []

      ws.append([instancia["NAME"]])
      ws.append([])

      ws.append(["Partículas:"] + nParticulas)
      ws.append(["Repetições:"] + nRepeticoes)
      ws.append(["Elite:"] + nElite)
      ws.append([])

      for i in range(0, len(dados)):
          ws.append(dados[i])
          dados[i] = dados[i][:1]

      ws.append([])
      wb.save(caminho_output)
      print(str(count) + "/" + str(len(arquivos_teste)), caso_teste)
      count += 1

print("Resultado salvo em", caminho_output)