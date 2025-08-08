import subprocess
import os
import pandas as pd
import numpy as np
import re
from openpyxl import load_workbook, Workbook
from sys import argv
from time import time
from leitor_instancia import ler_instancia

NUM_TESTES = 5
instance = ""

arquivos_teste = []
caminho = ''

if argv[1].upper() == "-F":
    arquivos_teste.append(argv[2])

elif argv[1].upper() == "-D":
    arquivos_teste = os.listdir(argv[2])

    caminho = argv[2]
    if caminho[-1] != '/':
        caminho += '/'
else:
    raise Exception("Não se sabe se é arquivo ou diretório!")

caminho_output = "./PSO_data.xlsx"

if len(argv) == 4:
    caminho_output = argv[3]

arquivos_teste.sort()

wb = None
ws = None
try:
    wb = load_workbook(caminho_output)
    ws = wb.active
    ws.append([])
except:
    wb = Workbook()
    ws = wb.active

dados = [["Mínimo:"],
         ["Mediana:"],
         ["Máximo:"],
         ["Média:"],
         ["Desvio padrão:"],
         ["Tempo:"]]

valores = []

nParticulas = [25]
nRepeticoes = [1000]
nElite = [5]

count = 1

for caso_teste in arquivos_teste:
    print(caso_teste)
    if caso_teste[-3:] != "vrp":
        continue

    instancia = ler_instancia(caminho + caso_teste)

    for c in range(len(nRepeticoes)):

        print("CONFIG", c)
        
        tempo = time()

        for i in range(1, NUM_TESTES +1):

            output = subprocess.check_output([
                            "./vrp_pso",
                            "-instancia",
                            caminho + caso_teste,
                            "-npart",
                            str(nParticulas[c]),
                            "-nrep",
                            str(nRepeticoes[c]),
                            "-seguir-melhor",
                            "10",
                            "-elite",
                            str(nElite[c]),
                            "-setorizar"
                            ]
                            ).decode()
            print(i)
            valores.append(float(output))
        
        tempo = (time() - tempo)/10

        dados[0].append(min(valores))
        dados[1].append(np.median(valores))
        dados[2].append(max(valores))
        dados[3].append(np.average(valores))
        dados[4].append(np.std(valores))
        dados[5].append(tempo)

        valores = []

    ws.append([instancia["NAME"]])
    ws.append([])

    ws.append(["Partículas:"] + nParticulas)
    ws.append(["Repetições:"] + nRepeticoes)
    ws.append(["Elite"] + nElite)
    ws.append([])

    #Colocar na tabela e resetar dados para partir para a próxima entrada
    for i in range(0, len(dados)):
        ws.append(dados[i])
        dados[i] = dados[i][:1]

    ws.append([])
    wb.save(caminho_output)
    print(str(count) + "/" + str(len(arquivos_teste)), caso_teste)
    count += 1

print("Resultado salvo em", caminho_output)